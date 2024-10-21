# Version 1.0
# Updated on 2024-08-01
# added elasticsearch functions(2024-08-01)

import requests
import logging, sys, io, os
from datetime import datetime, timezone, timedelta
from collections import Counter, defaultdict
import openpyxl
import base64
import jwt
import hashlib
import time
import re, subprocess
import calendar
from configparser import ConfigParser
import warnings, json
import deepsecurity
from deepsecurity.rest import ApiException
import jsonpickle, argparse
from elasticsearch import Elasticsearch
    
# Configure logging
logging.basicConfig(filename='vision_one.log', level=logging.INFO, 
                    format='%(asctime)s:%(levelname)s:%(message)s')

############################################## Config related functions #########################################

# config variable to handle config.ini file
config = ConfigParser()

def read_config():
    config_file = 'config.ini'
    
    if os.path.exists(config_file):
        if config.read(config_file):
            logging.info('config.ini file loaded')

            for section in config.sections():
                for key in config[section]:
                    if not config[section][key]:
                        logging.error(f"Configuration value for '{key}' in section '{section}' is empty. Please fill all the required fields.")
                        sys.exit(f"Configuration value for '{key}' in section '{section}' is empty. Please fill all the required fields.")
        else:
            logging.error("Error in reading the config.ini\nPlease check whether the correct config.ini is in the directory")
            sys.exit("Error in reading the config.ini\nPlease check whether the correct config.ini is in the directory")
    else:
        logging.warning("config.ini file is missing.\nCreating a new config.ini with default values.")

        if 'tenant_name' not in config.sections():
            config.add_section('tenant_name')
            config.set('tenant_name', 'a1_base_url', '')
            config.set('tenant_name', 'a1_application_id', '')
            config.set('tenant_name', 'a1_api_key', '')
            config.set('tenant_name', 'v1_base_url', "")
            config.set('tenant_name', 'v1_api_key', "")
            config.set('tenant_name', 'soar_base_url', '')
            config.set('tenant_name', 'soar_api_key', '')
            config.set('tenant_name', 'cloud_app_sec_base_url', '')
            config.set('tenant_name', 'cloud_app_sec_api_key', '')
            config.set('tenant_name', 'c1_base_url', "")
            config.set('tenant_name', 'c1_api_key', "")
            
            with open(config_file, 'w') as configfile:
                config.write(configfile)
            logging.info("Config File Created, Please fill the credentials.")
            sys.exit("Config File Created, Please fill the credentials.")

############################################### Elastic search configuration #########################################

def connect_elasticsearch():
    connection_check = False
    es = Elasticsearch(
        ["http://localhost:9200"],
        request_timeout=30,  # Use request_timeout instead of timeout
        max_retries=10,
        retry_on_timeout=True
    )
    if es.ping():
        logging.info("Elasticsearch connection established")
        connection_check = True
    else:
        logging.error("Elasticsearch connection Failed")
    return connection_check, es

def get_saved_reports_from_es():
    try:
        query = {
            "query": {
                "match_all": {}
            }
        }
        index = 'monthly_reports' if is_monthly_report else 'weekly_reports'
        query_with_size = {
            "size": 1000,
            **query  
        }
 
        response = es.search(index=index, body=query_with_size)
        hits = response['hits']['hits']
        return hits
    except Exception as e:
        logging.info(f"Not Getting Saved Reports from ElasticSearch: {e}")
        return []   

def get_id_of_es_file(file_name):
    try:
        is_file_present = False
        id = ''
        document = ''
        for hit in saved_reports:
            if file_name == hit['_source']['doc_name']:
                print(file_name, 'inside')
                print(hit['_source']['doc_name'])
                is_file_present = True
                id = hit['_id']
                document = hit['_source']['document']
                break
        return is_file_present, id, document
      
    except Exception as e:
        logging.error(f"Error Getting ID of Elasticsearch File: {e}")
        return False, None, None
                
############################################## Getting User Inputs ##############################################

def valid_date(date_string):
    try:
        return datetime.strptime(date_string, "%d/%m/%Y")
    except ValueError:
        try:
            return datetime.strptime(date_string, "%m/%Y")
        except ValueError:
            msg = f"Not a valid date: '{date_string}'. Expected formats: DD/MM/YYYY or MM/YYYY."
            raise argparse.ArgumentTypeError(msg)
            
def get_previous_three_months(month_name):
    try:
        month_number = list(calendar.month_name).index(month_name.capitalize())
        
        previous_months = []

        for _ in range(3):
            if month_number == 0:
                month_number = 12
            previous_months.append(month_number)
            month_number -= 1

        previous_months_names = [calendar.month_name[m] for m in reversed(previous_months)]

        return previous_months_names
    except Exception as e:
        logging.error(f"Error Fetching Past 3 months: {e}")

def get_days_in_month(year, month):
    return calendar.monthrange(year, month)[1]

def get_month_year(user_input):  
    try:
        if '/' in user_input:
            user_input_val = user_input.split('/')
            if len(user_input_val) == 2:
                monthly_report = True
            elif len(user_input_val) == 3:
                monthly_report = False
            else:
                raise ValueError("Invalid input format. Use MM/YYYY or DD/MM/YYYY.")
        elif '\\' in user_input:
            user_input_val = user_input.split('\\')
            if len(user_input_val) == 2:
                monthly_report = True
            elif len(user_input_val) == 3:
                monthly_report = False
            else:
                raise ValueError("Invalid input format. Use MM/YYYY or DD/MM/YYYY.")
        else:
            raise ValueError("Invalid input format. Use MM/YYYY or DD/MM/YYYY.")
        
        current_date = datetime.now()
        cr_year = current_date.year
        cr_month = current_date.month
        
        if not monthly_report:
            day, month, year = user_input_val
            day = int(day)
            cr_day = current_date.day     
        else:
            month, year = user_input_val
            
        month = int(month)
        year = int(year)
        req_days = get_days_in_month(year, month)
        if not monthly_report:
            if day < 1 or day > req_days:
                raise ValueError(f"Date must be between 1 and {req_days}.")
            if year == cr_year and month == cr_month and day > cr_day:
                raise ValueError("Invalid day.")
            
        if month < 1 or month > 12:
            raise ValueError("Month must be between 1 and 12.")
        if year < 1000:
            raise ValueError("Year must be a four-digit number.")
        if year > cr_year:
            raise ValueError("Invalid year.")
        if year == cr_year and month > cr_month:
            raise ValueError("Invalid month.")
        
        if monthly_report:
            logging.info(f"Received valid month and year: {month}/{year}")
            return monthly_report, month, year, req_days
        else:
            logging.info(f"Received valid date, month, and year: {day}/{month}/{year}")
            return monthly_report, month, year, day
    except ValueError as e:
        logging.error(f"Error: {e}")
        sys.exit(f"Invalid input: {e}")

def get_tenant_name():
    try:
        tenant_name_input = args.tenants
        config_tenant_names = [section for section in config.sections()]
        config_tenant_names_in_lower = [section.lower() for section in config.sections()]
        
        tenant_name_map = {name.lower(): name for name in config_tenant_names}
        
        if 'all' in [name.lower() for name in tenant_name_input]:
            logging.info("You selected: All tenants")
            return config_tenant_names
        
        selected_tenants = []
        invalid_names = []
        for name in tenant_name_input:
            if name.lower() in tenant_name_map:
                selected_tenants.append(tenant_name_map[name.lower()])
            else:
                invalid_names.append(name)
        
        if invalid_names:
            raise ValueError(f"Invalid tenant name(s): {', '.join(invalid_names)}. Please provide valid tenant name(s).")
        
        logging.info(f"You selected: {', '.join(selected_tenants)}")
        return selected_tenants
    except ValueError as e:
        logging.error(f"Error: {e}")
        sys.exit(f"Invalid input: {e}")

def get_start_and_end_date(original_date):
    try:
        user_input = original_date

        date_details = get_month_year(user_input)
        year = date_details[2]
        month = date_details[1]
        month = str(month).zfill(2)
        month_name = calendar.month_name[int(month)]
        
        if date_details[0]:
            days = date_details[3]
            start_date = f'{year}-{str(month)}-01T00:00:00Z'
            end_date = f'{year}-{str(month)}-{days}T23:59:00Z'
        else:
            day = date_details[3]
            day = str(day).zfill(2)
            end_date_str = f'{year}-{month}-{day}T00:00:00Z'
            end_date = datetime.strptime(end_date_str, '%Y-%m-%dT%H:%M:%SZ')
            start_date = end_date - timedelta(weeks=1)
            end_date = end_date.strftime('%Y-%m-%dT%H:%M:%SZ')
            start_date = start_date.strftime('%Y-%m-%dT%H:%M:%SZ')
        return date_details[0], start_date, end_date, month_name, date_details[3]
    except Exception as e:
        logging.error(f"Error Getting Start and End Date: {e}")
                  
############################################# Cloud App Security Functions #########################################

def make_api_call_cloud_app_security(url, params, value):
    url = cloud_app_sec_base_url + url
    headers = {'Authorization': f'Bearer {cloud_app_sec_api_key}'}
    logging.info(f"Making api call to Cloud App Security...")
    
    params.update({
        'start': start_date,
        'limit': 1000
    })
    
    total_events = []
    try:
        response = requests.get(url, headers=headers, params=params)
        response.raise_for_status()
        response = response.json()       
        total_events.extend(response.get(value, []))
        next_link = response.get('next_link', '')
        while next_link:
            response = requests.get(url=next_link, headers=headers)
            response.raise_for_status()
            response = response.json()
            total_events.extend(response.get(value, []))
            next_link = response.get('next_link', '')          
        
        return total_events
    except requests.exceptions.RequestException as e:
        response.json()
        error_msg = response['message'] if response['message'] else response['msg']
        logging.error(f"Error Making api call to Cloud App Security: {e}, error message: {error_msg}")
        return []
    except Exception as e:
        logging.error(f"Unexpected Error Making api call to Cloud App Security: {e}")
        return []

def get_quarantine_events():
    quarantine_events = make_api_call_cloud_app_security('/v1/siem/quarantine_events', {'service': 'exchange'}, 'quarantine_events')
    logging.info('Fetching quarantine events...')
    data = {
            'email_status': {'Quarantined': 0, 'Deleted': 0},
            'threat_type': {'Malicious Files': 0, 'Malicious URLs': 0, 'Phishing': 0, 'Spoofing': 0, 'Suspicious Object': 0, 'Blocked Object': 0}
        }
    try:       
        for event in quarantine_events:
            event = event['message']
            if event['mail_status'] == 'Quarantined':
                data['email_status']['Quarantined'] += 1
            elif event['mail_status'] == 'Deleted':
                data['email_status']['Deleted'] += 1
            if event['threat_type'] == 'Malicious Files':
                data['threat_type']['Malicious Files'] += 1
            elif event['threat_type'] == 'Malicious URLs':
                data['threat_type']['Malicious URLs'] += 1
            elif event['threat_type'] == 'Phishing':
                data['threat_type']['Phishing'] += 1
            elif event['threat_type'] == 'Spoofing':
                data['threat_type']['Spoofing'] += 1
            elif event['threat_type'] == 'Suspicious Object':
                data['threat_type']['Suspicious Object'] += 1
            elif event['threat_type'] == 'Blocked Object':
                data['threat_type']['Blocked Object'] += 1
        return data
    except Exception as e:
        logging.error(f"Unexpected Error fetching Quarantine Events: {e}")
        return {}

def get_sweeping_mails():
    try:
        sweeping_mails = make_api_call_cloud_app_security('/v1/sweeping/mails', {'end': end_date}, 'value')
        logging.info('Fetching Sweeping Mails...')
        
        senders_list = []
        receipts_list = []
        
        for sweeping_mail in sweeping_mails:
            senders_list.append(sweeping_mail['mail_message_sender'])
            receipts_list.extend(sweeping_mail['mail_message_recipient'])
        
        top3_senders = [sender for sender, _ in Counter(senders_list).most_common(3)]
        top3_receipts = [receipt for receipt, _ in Counter(receipts_list).most_common(3)]
        
        return {
            'top3_senders': top3_senders,
            'top3_receipts': top3_receipts
        }
        
    except Exception as e:
        logging.error(f"Unexpected Error fetching Sweeping Mails: {e}")
        return {}

############################################# Cloud One Functions #########################################

def make_cloud_one_api_call(api, method):
    if not sys.warnoptions:
        warnings.simplefilter("ignore")
    configuration = deepsecurity.Configuration()
    configuration.host = c1_base_url
    
    logging.info("Making api call to Cloud One...")
    
    # Authentication
    configuration.api_key['api-secret-key'] = c1_token
    configuration.api_key['Authorization'] = c1_token

    # Initialization
    # Set Any Required Values
    api_class = getattr(deepsecurity, api)
    api_instance = api_class(deepsecurity.ApiClient(configuration))
    api_version = 'v1'
    search_filter = deepsecurity.SearchFilter()
    search_criteria = deepsecurity.SearchCriteria()
    search_criteria.id_value = 42
    search_filter.search_criteria = [search_criteria]


    try:
        api_method = getattr(api_instance, method)
        api_response = api_method(api_version)
        api_response = jsonpickle.encode(api_response, unpicklable=False)
        api_response = json.loads(api_response)
        return api_response
    except ApiException as e:
        logging.error("An exception occurred when making api call to Cloud One: %s\n" % e)
        return {}
    except Exception as e:
        logging.error("Unexpected error occured when making api call to Cloud One: %s\n" % e)
        return {}
    
def key_feature_adoption_rate_of_c1():
    c1_response = make_cloud_one_api_call("ComputersApi", "list_computers")
    try:
        list_computers = c1_response.get('_computers', [])
        result_data  = {
            "Intrusion Prevention System (IPS)": {'total': 0, 'count': 0},
            "Anti-malware": {'total': 0, 'count': 0},
            "Web Reputation": {'total': 0, 'count': 0},
            "Behavior Monitoring": {'total': 0, 'count': 0},
            "Predictive Machine Learning": {'total': 0, 'count': 0},
            "Smart Feedback": {'total': 0, 'count': 0},
            "Firewall": {'total': 0, 'count': 0},
            "Agent Self-Protection": {'total': 0, 'count': 0},
            "File Integrity Monitoring (FIM)": {'total': 0, 'count': 0},
            "Log Inspection": {'total': 0, 'count': 0},
            "Application Control": {'total': 0, 'count': 0}                               
        }
        
        looping_data = [
            ["Intrusion Prevention System (IPS)", 'Yes', 'No'],
            ["Anti-malware", 'on', 'off'],
            ["Web Reputation", 'on', 'off'],
            ["Behavior Monitoring", '', ''],
            ["Predictive Machine Learning", '', ''],
            ["Smart Feedback", 'true', 'false'],    
            ["Firewall", 'on', 'off'],
            ["Agent Self-Protection", 'true', 'false'],
            ["File Integrity Monitoring (FIM)", 'real-time', 'off'],
            ["Log Inspection", 'on', 'off'],
            ["Application Control", 'on', 'off']
        ]
        for data in list_computers:
            for loop in looping_data:
                if loop[0] == "Intrusion Prevention System (IPS)":
                    value = data.get('_computer_settings', {}).get('_intrusion_prevention_setting_auto_apply_recommendations_enabled', {}).get('_value', '')
                elif loop[0] == "Anti-malware":
                    value = data.get('_anti_malware', {}).get('_state', '')
                elif loop[0] == "Web Reputation":
                    value = data.get('_web_reputation', {}).get('_state', '')
                elif loop[0] == "Behavior Monitoring":
                    value = data.get('_computer_settings', {}).get('_anti_malware_setting_behavior_monitoring_scan_exclusion_list', {}).get('_value', '')
                elif loop[0] == "Predictive Machine Learning":
                    value = data.get('_computer_settings', {}).get('anti_malware_setting_predictive_machine_learning_exceptions', {}).get('_value', '')
                elif loop[0] == "Smart Feedback":
                    value = data.get('_computer_settings', {}).get('_anti_malware_setting_smart_protection_global_server_enabled', {}).get('_value', '')
                elif loop[0] == "Firewall":
                    value = data.get('_firewall', {}).get('_state', '')
                elif loop[0] == "Agent Self-Protection":
                    value = data.get('_computer_settings', {}).get('_platform_setting_agent_self_protection_enabled', {}).get('_value', '')
                elif loop[0] == "File Integrity Monitoring (FIM)":
                    value = data.get('_integrity_monitoring', {}).get('_state', '')
                elif loop[0] == "Log Inspection":
                    value = data.get('_log_inspection', {}).get('_state', '')
                elif loop[0] == "Application Control":
                    value = data.get('_application_control', {}).get('_state', '')
                
                if value == loop[1]:
                    result_data[loop[0]]['total'] += 1 
                    result_data[loop[0]]['count'] += 1 
                elif value == loop[2]:
                    result_data[loop[0]]['total'] += 1
        
        return result_data            
                
    except Exception as e:
        logging.error(f"Error Fetching C1 Data: {e}")
        return {}
                                   
############################################# Apex One Funtions #########################################

def create_checksum(http_method, raw_url, headers, request_body):
    try:
        string_to_hash = http_method.upper() + '|' + raw_url.lower() + '|' + headers + '|' + request_body
        base64_string = base64.b64encode(hashlib.sha256(str.encode(string_to_hash)).digest()).decode('utf-8')
    except Exception as e:
        logging.error(f"Error Create Checksum for ApexOne: {e}")
        return
    return base64_string

def create_jwt_token(appication_id, api_key, http_method, raw_url, headers, request_body,
                     iat=time.time(), algorithm='HS256', version='V1'):
    try:
        payload = {'appid': appication_id,
                'iat': iat,
                'version': version,
                'checksum': create_checksum(http_method, raw_url, headers, request_body)}
        token = jwt.encode(payload, api_key, algorithm=algorithm)
    except Exception as e:
        logging.error(f"Error Create JWT TOken for ApexOne: {e}")
        return
    return token 

def make_apex_one_api_call(url):
    try:
        dt = datetime.fromisoformat(start_date[:-1])
        unix_timestamp = int(dt.replace(tzinfo=timezone.utc).timestamp())
        
        productAgentAPIPath = url
        canonicalRequestHeaders = ''

        useRequestBody = ''
        useQueryString = f"?output_format=CEF&page_token=0&since_time={unix_timestamp}"

        total_data = []
        full_url = a1_base_url + productAgentAPIPath + useQueryString        
                   
        while full_url:
            jwt_token = create_jwt_token(a1_application_id, a1_api_key, 'GET',
                                    productAgentAPIPath + useQueryString,
                                    canonicalRequestHeaders, useRequestBody, iat=time.time())
        
            headers = {'Authorization': 'Bearer ' + jwt_token , 'Content-Type': 'application/json;charset=utf-8'}
            response = requests.get(full_url, headers=headers, verify=False)
            response = response.json()
            logs = response.get("Data", {}).get("Logs", [])
            nextlink = response.get("Data", {}).get("Next", None)
            if logs:
                total_data.extend(logs)

            if not nextlink:
                break
            useQueryString = nextlink[nextlink.index('?'):]
            full_url = a1_base_url + productAgentAPIPath + useQueryString 
            
    except requests.exceptions.RequestException as e:
        logging.error(f"Error fetching Apex One data: {e}, error message: {response.json()['Meta']['ErrorMsg']}")
        return []
    except Exception as e:
        logging.error(f"Unexpected error occurred while making Apex One API Call: {e}")
        return []
    return total_data

def get_syslog_data(log_type, action_type_list: list[str]):
    data = make_apex_one_api_call(f"/WebApp/api/v1/logs/{log_type}")
    
    try:
        return_data = {
            "total_count": [],
            "top_3_endpoints": []
        }
        
        pattern = re.compile(r"\b\w*dhost\w*=(\S+)", re.IGNORECASE)
        
        for action_type in action_type_list:
            count = 0
            action_type_logs = {}
            
            for log in data:
                if action_type in log:
                    count += 1
                    match = pattern.search(log)
                    if match:
                        dhost_value = match.group(1)
                        action_type_logs[dhost_value] = action_type_logs.get(dhost_value, 0) + 1
            
            return_data["total_count"].append({action_type: count})
            
            top_3_endpoints = sorted(action_type_logs.items(), key=lambda x: x[1], reverse=True)[:3]
            return_data["top_3_endpoints"].append({action_type: top_3_endpoints})
            
    except Exception as e:
        logging.error(f"Unexpected error occurred fetching Syslog Data: {e}")
        return []
    
    return return_data

############################################## SOAR Functions ###############################################

def get_soar_alerts(api_identifier):
    try:
        url = f"{soar_base_url}/api/query_builder/widget/?api_identifier={api_identifier}"

        payload = json.dumps({
        "query": [
            {
            "api_key": soar_api_key,
            "from_date": start_date,
            "to_date": end_date
            }
        ]
        })
        soar_headers = {
            'Content-Type': 'application/json'
        }

        response = requests.request("POST", url, headers=soar_headers, data=payload)
        response = response.json()
    except Exception as e:
        logging.error(f"Error fetching while making API call to SOAR for {api_identifier}: {e}")
        return []
    return response    

def get_incident_categories(data):
    try:
        categories = set()
        for alerts in data:
            for alert in alerts["query_result"]:
                category = alert.get("category")
                categories.add(category)
    except Exception as e:
        logging.error(f"Error fetching Alerts for getting incidents by category: {e}")
        return []
    return list(categories)   

def get_incident_summary_alerts(api_identifier, key, values_list):
    data = get_soar_alerts(api_identifier)
    try:
        metrics = {
            "closed_with_resolution": {v: 0 for v in values_list},
            "closed_without_acknowledgement": {v: 0 for v in values_list},
            "pending_with_soc_team": {v: 0 for v in values_list}
        }

        def update_metrics(value, closure_code, count):
            if closure_code == "Lack of Information":
                metrics["closed_without_acknowledgement"][value] += count
            elif closure_code == "Waiting":
                metrics["pending_with_soc_team"][value] += count
            elif closure_code not in ["Not Set", "Lack of Information", "Waiting"]:
                metrics["closed_with_resolution"][value] += count

        for alerts in data:
            for alert in alerts["query_result"]:
                value = alert.get(key)
                closure_code = alert.get("closure_code")
                if alert.get("count"):
                    count = alert.get("count")
                else:
                    count = alert.get("SLA_Not_Met", 0) + alert.get("SLA_Met", 0)

                if value is None or closure_code is None or count is None:
                    logging.error(f"Missing key in alert: {alert}")
                    continue

                if value in values_list:
                    update_metrics(value, closure_code, count)

    except Exception as e:
        logging.error(f"Error fetching Alerts for {api_identifier}: {e}")
        return []

    return metrics
    
def pending_incidents():
    data = get_soar_alerts("pending-incidents")
    try:
        incident_dict = {}
        for alerts in data:
            for alert in alerts["query_result"]:
                if '] ' in alert["title"]:
                    name = alert["title"].split('] ', 1)[1]
                else:
                    name = alert["title"]
                if name in incident_dict:
                    incident_dict[name]["count"] += 1
                else:
                    incident_dict[name] = {"title": name, "count": 1, "priority": alert["priority"]}
    except Exception as e:
        logging.error(f"Error fetching Pending Incidents: {e}")
        return []
    
    return list(incident_dict.values())           
        
############################################## Vision One API functions #########################################

def make_v1_api_call(url_suffix, params, headers1):
    url = v1_base_url + url_suffix
    headers1.update(headers)
    total_data = []
    logging.info(f"Fetching data from {url_suffix} endpoint")
    try:
        response = requests.get(url=url, headers=headers1, params=params)
        response.raise_for_status()
        response = response.json()
        total_data.extend(response['items'])
        
        while "nextLink" in response:
            url = response['nextLink']
            response = requests.get(url=url, headers=headers1)
            response.raise_for_status()
            response = response.json()
            total_data.extend(response['items'])
            
    except requests.exceptions.RequestException as e:
        logging.error(f"Error fetching {url_suffix} data: {e}, error message: {response.json().get('error', {}).get('message', '')}")
        return []
    except Exception as e:
        logging.error(f"Unexpected error fetching {url_suffix} data: {e}")
        return []
    
    return total_data

def isvalid_token():
    try:
        url_suffix = "/v3.0/healthcheck/connectivity"
        url = v1_base_url + url_suffix
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        response = response.json()
        logging.info(f"Valid API token and URL: {response}")
    except requests.exceptions.RequestException as e:
        logging.error(f"Error fetching base url: {e}, error message: {response.json()['error']['message']}")
        sys.exit(1)
    return response
    
def incident_overview():
    try:
        incidents = []

        logging.info("Fetching Incident Overview")

        for workbench_alert in wb_data:
            created_datetime = workbench_alert.get('createdDateTime', '')
            datetime_object = datetime.strptime(created_datetime, "%Y-%m-%dT%H:%M:%SZ")

            wb_month = datetime_object.month
            wb_year = datetime_object.year

            matched_rules_names = ', '.join(rule['name'] for rule in workbench_alert.get('matchedRules', []))

            incident = {
                "workbench_id": workbench_alert.get('id', ''),
                "description": workbench_alert.get('description', None),
                "createdDateTime": created_datetime,
                "month&Year": f"{wb_month}/{wb_year}",
                "name": matched_rules_names,
            }

            incidents.append(incident)

    except Exception as e:
        logging.error(f"Error fetching Incident Overview: {e}")
        return None
    
    return incidents

def critical_alerts():
    critical_alerts = make_v1_api_call("/v3.0/workbench/alerts", {"top": 1000, "startDateTime": start_date, "endDateTime": end_date}, {"TMV1-Filter": "severity eq 'high' or severity eq 'critical'"})
    try:
        critical_high_alerts = []        
        for alert in critical_alerts:
            name = ''
            count = 0
            if "matchedRules" in alert:
                for x in alert['matchedRules']:
                    name += x['name'] + ', '
                    count += 1
            name = name[:-2]

            existing_alert = next((item for item in critical_high_alerts if item['name'] == name), None)
            
            if existing_alert:
                existing_alert['count'] += count
            else:
                data = {
                    "severity": alert['severity'],
                    "name": name,
                    "count": count
                }
                critical_high_alerts.append(data)
               
    except Exception as e:
        logging.error(f"Error Fetching Critical / High Alerts: {e}")
        return       
    return critical_high_alerts

def top_5_alerts():
    try:
        logging.info("Fetching Top 5 Alerts")

        incidents = [
            {
                "workbench_id": alert['id'],
                "description": alert.get('description'),
                "createdDateTime": alert['createdDateTime'],
                "month&Year": f"{datetime.strptime(alert['createdDateTime'], '%Y-%m-%dT%H:%M:%SZ').month}/{datetime.strptime(alert['createdDateTime'], '%Y-%m-%dT%H:%M:%SZ').year}",
                "alertProvider": alert['alertProvider'],
                "severity": alert['severity'],
                "name": ', '.join(rule['name'] for rule in alert.get('matchedRules', []))
            }
            for alert in wb_data
        ]

        top_alerts = (Counter(alert['name'] for alert in incidents)).most_common(5)

        top_alert_details = []
        for alert_name, count in top_alerts:
            details = next(incident for incident in incidents if incident['name'] == alert_name)
            details['NoofOccurences'] = count
            top_alert_details.append(details)

    except Exception as e:
        logging.error(f"Error Fetching Top 5 Alerts: {e}")
        return None

    return top_alert_details

def get_security_posture():
    url_suffix = "/v3.0/asrm/securityPosture"
    url = v1_base_url + url_suffix
    logging.info("Fetching Security Posture of the Organization")
    try:
        riskIndex = 0
        highlyExpCves = 0
        vulnerableAssesment = {}
        securityConfig = {}
        stdEndPoint = {}
        serverWorkload = {}
        
        response = requests.get(url=url, headers=headers)
        response.raise_for_status()
        response = response.json()
        
        logging.info("Fetching Risk Index and Risk Level")
        if response['riskIndex'] >= 0 and response['riskIndex'] <= 30:
            riskLevel = 'Low'
        elif response['riskIndex'] >= 31 and response['riskIndex'] <= 70:
            riskLevel = 'Medium'
        elif response['riskIndex'] >= 71 and response['riskIndex'] <= 100:
            riskLevel = 'High'
        riskIndex = response['riskIndex']
        
        logging.info('Fetching Highly Exploitable CVEs')
        highlyExpCves = response['cveManagementMetrics']['count']
        
        logging.info('Fetching Vulnerability Assessment Report')
        vulnerableAssesment = response['cveManagementMetrics']
        
        logging.info('Fetching Security Configuration Report')
        securityConfig = response['exposureStatus']['domainAccountMisconfigurationStatus']
        
        logging.info('Fetching Agent Life Cycle Report')
        agentLifeCycle = response['securityConfigurationStatus']['endpointAgentStatus']['agentVersionStatus']
        
        logging.info('Fetching Adoption Rate of Standard Endpoint Protection')
        stdEndPoint = response['securityConfigurationStatus']['endpointAgentStatus']['agentFeatureStatus']['standardEndpointProtection']
        
        logging.info('Fetching Adoption Rate of Server Workload Protection')
        serverWorkload = response['securityConfigurationStatus']['endpointAgentStatus']['agentFeatureStatus']['serverWorkloadProtection']
        
        insecure_host_count = response['exposureStatus']['insecureHostConnectionStatus']['connectionIssueCount']
        data = {
            'Risk Index': {
                'riskIndex': riskIndex,
                'riskLevel': riskLevel
            },
            'Highly Exploitable Unique CVES': highlyExpCves,
            'Vulnerability Assesment': vulnerableAssesment,
            'Security Configuration': securityConfig,
            'Agent Life Cycle': agentLifeCycle,
            'Adoption Rate of Standard Endpoint Protection': stdEndPoint,
            'Adoption Rate of Server Workload Protection': serverWorkload,
            'Insecure Host Count': insecure_host_count
        }
        
    except requests.exceptions.RequestException as e:
        logging.error(f"Error fetching security posture: {e}, error message: {response.json()['error']['message']}")
        return {}
    except Exception as e:
        logging.error(f"Unexpected Error fetching security posture: {e}")
        return {}
    return data

def save_to_file(name: str, data: dict):
    try:
        with open(f"{name}.json", 'w') as f:
            f.write(json.dumps(data))
        logging.info(f"Data saved to {name}.json")
    except Exception as e:
        logging.error(f"Error saving data to file: {e}")

def get_top_risk_users():
    url_suffix = "/v3.0/asrm/highRiskUsers"
    url = v1_base_url + url_suffix
    params = {
        "top": 10
    }
    logging.info("Fetching top risk users")
    try:
        response = requests.get(url=url, headers=headers, params=params)
        response.raise_for_status()
        response = response.json()
        data = []
        for user in response["items"]:
            data.append({
                "userName": user["userName"],
                "riskScore": user["riskScore"]
            })
        
    except requests.exceptions.RequestException as e:
        logging.error(f"Error fetching top risk users: {e}, error message: {response.json()['error']['message']}")
        return []
    except Exception as e:
        logging.error(f"Unexpected Error fetching top risk users: {e}")
        return []
    return data

def get_top_risk_devices():
    url_suffix = "/v3.0/asrm/highRiskDevices"
    url = v1_base_url + url_suffix
    params = {
        "top": 10
    }
    logging.info("Fetching top risk devices")
    try:
        response = requests.get(url=url, headers=headers, params=params)
        response.raise_for_status()
        response = response.json()
        data = []
        for device in response["items"]:
            data.append({
                "deviceName": device["deviceName"],
                "riskScore": device["riskScore"]
            })

    except requests.exceptions.RequestException as e:
        logging.error(f"Error fetching top risk devices: {e}, error message: {response.json()['error']['message']}")
        return []
    except Exception as e:
        logging.error(f"Unexpected Error fetching top risk devices: {e}")
        return []
    return data

def get_account_compromise_events():
    url_suffix = "/v3.0/asrm/accountCompromiseIndicators"
    url = v1_base_url + url_suffix
    logging.info("Fetching account compromise events")
    total_events = []
    headers1 = {
        "TMV1-Filter": "riskLevel eq 'high'"
    }
    headers1.update(headers)
    try:
        response = requests.get(url=url, headers=headers1)
        logging.info(response.status_code)
        response.raise_for_status()
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
        response = response.json()
        total_events.extend(response['items'])

        while "nextLink" in response:
            url = response['nextLink']
            response = requests.get(url=url, headers=headers1)
            response.raise_for_status()
            response = response.json()
            total_events.extend(response['items'])
        
        data = []
        for event in total_events:
            data.append({
                "account": event["account"],
                "riskLevel": event["riskEvent"]["riskLevel"],
                "dataSourceOrProcessor": event["dataSourceOrProcessor"],
                "riskEventName": event["riskEvent"]["name"]
            })
        
    except requests.exceptions.RequestException as e:
        logging.error(f"Error Fetching Account Compromise Events: {e}, error message: {response.json()['error']['message']}")
        return []
    except Exception as e:
        logging.error(f"Unexpected Error Fetching Account Compromise Events: {e}")
        return []
    
    return data

def get_endpoint_details(filter):
    url_suffix = "/v3.0/endpointSecurity/endpoints" 
    url = v1_base_url + url_suffix
    logging.info("Fetching Endpoint Feature Compliance details")
    headers1 = {
        "TMV1-Filter": filter
    }   
    headers1.update(headers)
    try:
        response = requests.get(url=url, headers=headers1)
        response.raise_for_status()
        response = response.json()
        
    except requests.exceptions.RequestException as e:
        logging.error(f"Error Fetching Endpoint Feature Compliance details: {e}, error message: {response.json()['error']['message']}")
        return []
    except Exception as e:
        logging.error(f"Unexpected Error Fetching Endpoint Feature Compliance details: {e}")
        return []
    return response["totalCount"]

def get_component_versions():
    try:
        visionOne_data["Component Versions"] = {
            "WorkLoad Protection" : {
                "Older": get_endpoint_details("eppAgentComponentVersion eq outdatedVersion and securityDeployment eq workloadProtection"),
                "Latest": get_endpoint_details("eppAgentComponentVersion eq latestVersion and securityDeployment eq workloadProtection"),
                "End of Life": get_endpoint_details("eppAgentComponentVersion eq unknownVersions and securityDeployment eq workloadProtection")
            },
            "Standard Endpoint Protection" : {
                "Older": get_endpoint_details("eppAgentComponentVersion eq outdatedVersion and securityDeployment eq userProtection"),
                "Latest": get_endpoint_details("eppAgentComponentVersion eq latestVersion and securityDeployment eq userProtection"),
                "End of Life": get_endpoint_details("eppAgentComponentVersion eq unknownVersions and securityDeployment eq userProtection")
            }
        }
    except Exception as e:
        logging.error(f"Unexpected Error fetching Component Versions: {e}")
        return {}

def get_top10_cves():
    url_suffix = '/v3.0/asrm/vulnerableDevices'
    headers1 = {
        'TMV1-Filter': 'cveEventRiskLevel eq "high" or cveEventRiskLevel eq "medium"',
    }
    headers1.update(headers)
    url = v1_base_url + url_suffix

    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        
        vulnerable_devices = response.json().get('items', [])
        unique_cve_ids = {}

        for device in vulnerable_devices:
            cve_record = device['cveRecords'][0]
            cve_id = cve_record['id']
            event_risk_level = cve_record['eventRiskLevel']
            if event_risk_level in ['high', 'medium']:
                if cve_id not in unique_cve_ids:
                    unique_cve_ids[cve_id] = cve_record['cvssScore']

        top_10_unique_cves = [{'cve_id': cve_id, 'score': score} for cve_id, score in unique_cve_ids.items()]
        return top_10_unique_cves[:10]


    except requests.exceptions.RequestException as e:
        logging.error(f'Error Fetching Top Vulnerable CVEs: {e}, error message: {response.text}')
        return []
    except Exception as e:
        logging.error(f"Unexpected Error Fetching Top Vulnerable CVEs: {e}")
        return []

def get_top5_endpoints():
    total_data = make_v1_api_call("/v3.0/oat/detections", {"detectedStartDateTime": start_date, "detectedEndDateTime": end_date, "top": 200}, {})
    try:
        aggregated_data = defaultdict(lambda: {'count': 0, 'riskLevels': defaultdict(int)})
        
        for item in total_data:
            endpoinName = item.get('endpoint', {}).get('endpointName', 'N/A')
            riskLevels = [filter.get('riskLevel', 'N/A') for filter in item.get('filters', [])]
            
            aggregated_data[endpoinName]['count'] += 1
            for riskLevel in riskLevels:
                aggregated_data[endpoinName]['riskLevels'][riskLevel] += 1
        
        output_data = sorted(
            [{'endpoinName': endpoint, 'count': data['count'], 'riskLevels': dict(data['riskLevels'])} 
            for endpoint, data in aggregated_data.items()],
            key=lambda x: x['count'], 
            reverse=True
        )[:5]
    except Exception as e:
        logging.error(f"Error Generating Top 5 Endpoints: {e}")
        return []
    
    return output_data
   
################################################ Monthly Excel format Functions #################################################

def xl_top_risk_users():
    data = get_top_risk_users()
    try:
        risk_user_sheet = workbook.create_sheet(title="Top Risk Users")
        db_data = {"Top Risk Users": [] }
        risk_users_data = [
            ["Name of the Users", "Latest Risk Score"]
        ]

        for users in data:
            risk_users_data.append([users["userName"], users["riskScore"]])
            db_data["Top Risk Users"].append({"User Name": users["userName"], "Risk Score": users["riskScore"]})

        visionOne_data["Top Risk Users"] = db_data
        
        risk_user_sheet.merge_cells('A1:B1')
        risk_user_sheet['A1'] = "Top Risk Users"
        risk_user_sheet['D1'] = "Recommendation/Note/Summary"

        for row in risk_users_data:
            risk_user_sheet.append(row)
    except Exception as e:
        logging.error(f"Error Adding data to Top Risk Users: {e}")
        return
    
def xl_top_risk_devices():
    data = get_top_risk_devices()
    try:
        risk_device_sheet = workbook.create_sheet(title="Top Risk Device")
        db_data = {"Top Risk Device": [] }
        risk_devices_data = [
            ["Name of the Devices", "Latest Risk Score"]
        ]

        for devices in data:
            risk_devices_data.append([devices["deviceName"], devices["riskScore"]])
            db_data["Top Risk Device"].append({"Device Name": devices["deviceName"], "Risk Score": devices["riskScore"]})

        visionOne_data["Top Risk Device"] = db_data
            
        risk_device_sheet.merge_cells('A1:B1')
        risk_device_sheet['A1'] = "Top Risk Device"
        risk_device_sheet['D1'] = "Recommendation/Note/Summary"

        for row in risk_devices_data:
            risk_device_sheet.append(row)
    except Exception as e:
        logging.error(f"Error Adding data to Top Risk Device: {e}")
        return

def xl_account_compromise_events():
    data = get_account_compromise_events()
    try:
        account_compromise_events_sheet = workbook.create_sheet(title="Account Compromise Events")
        
        db_data = {"Account Compromise Events": []}
        account_compromise_events_data = [
            ["Risk Event", "Data Source", "Asset", "Event Risk"]
        ]

        for event in data:
            account_compromise_events_data.append([event["riskEventName"], event["dataSourceOrProcessor"], event["account"], event["riskLevel"]])
            db_data["Account Compromise Events"].append({"Risk Event":event["riskEventName"], "Data Source":event["dataSourceOrProcessor"], "Asset":event["account"], "Event Risk":event["riskLevel"]})
           
        visionOne_data["Account Compromise Events"] = db_data
            
        account_compromise_events_sheet.merge_cells('A1:D1')
        account_compromise_events_sheet['A1'] = "Account Compromise Events"
        account_compromise_events_sheet['F1'] = "Recommendation/Note/Summary"

        for row in account_compromise_events_data:
            account_compromise_events_sheet.append(row)
    except Exception as e:
        logging.error(f"Error Adding data to Account Compromise Events: {e}")
        return

def xl_critical_high_incidents_summary():
    data = critical_alerts()
    try:
        critical_alerts_sheet = workbook.create_sheet(title="Critical High Incidents Summary")
        
        db_data = {"Critical/High Incidents Summary": []}
        critical_alerts_data = [
            ["Impact -Priority", "Incident Title", "Findings and Action Performed"]
        ]

        for alert in data:
            priority = 'P1 - Critical' if alert["severity"] == "critical" else 'P2 - High'
            title = alert["name"] + " - " + str(alert["count"])
            critical_alerts_data.append([priority, title])
            db_data["Critical/High Incidents Summary"].append({"Impact-Priority": priority, "Incident Title": title, "Findings and Actions Performed": ""})
        
        visionOne_data["Critical High Incidents Summary"] = db_data
           
        critical_alerts_sheet.merge_cells('A1:C1')
        critical_alerts_sheet['A1'] = "Critical/High Incidents Summary"

        for row in critical_alerts_data:
            critical_alerts_sheet.append(row)
    except Exception as e:
        logging.error(f"Error Adding data to Critical High Incidents Summary: {e}")
        return
  
def xl_executive_summary():
    data = top_5_alerts()
    try:
        executive_summary_sheet = workbook.create_sheet(title="Executive Summary")
        
        db_data = {"Incident Overview by SOC Team": [], "Action Performed by SOC Team": [], "Recommendations by SOC Team": []}
        executive_summary_data = [
            ["No of Incidents", "Name of Incident", "Description", "", "No of Incidents", "Name of Incident", "Action Performed", "", "No of Incidents", "Name of Incident", "Recommendations"]
        ]

        for alert in data:
            count = alert["NoofOccurences"]
            name = alert["name"]
            description = alert["description"]

            executive_summary_data.append([count, name, description, "", count, name, "", "", count, name])
            db_data["Incident Overview by SOC Team"].append({"No of Incidents": count, "Name of Incident": name, "Description": description})
            db_data["Action Performed by SOC Team"].append({"No of Incidents": count, "Name of Incident": name, "Action Performed": ""})
            db_data["Recommendations by SOC Team"].append({"No of Incidents": count, "Name of Incident": name, "Recommendations": ""})
            
        visionOne_data["Executive Summary"] = db_data
            
        executive_summary_sheet.merge_cells('A1:C1')
        executive_summary_sheet['A1'] = "Incident Overview By SOC Team"

        executive_summary_sheet.merge_cells('E1:G1')
        executive_summary_sheet['E1'] = "Action Performed by SOCTeam"

        executive_summary_sheet.merge_cells('I1:K1')
        executive_summary_sheet['I1'] = "Recommendations by SOC Team"

        executive_summary_sheet['M1'] = "Conclusion"

        for row in executive_summary_data:
            executive_summary_sheet.append(row)
    except Exception as e:
        logging.error(f"Error Adding data to Executive summary: {e}")
        return
    
def xl_detailed_summary():
    incident = incident_overview()
    top_5_alert = top_5_alerts()
    try:
        detailed_summary_sheet = workbook.create_sheet(title="Detailed Summary")
        count = security_posture_data.get("Highly Exploitable Unique CVES", 'N/A')
        riskIndex = security_posture_data.get('Risk Index', {}).get("riskIndex", 'N/A')
        description = ("The Risk Index is a comprehensive score based on the dynamic assessment of risk factors inclusive "
                       "of exposure, attack risk, and security configurations risk. You can take company-wide remediation steps "
                       "and preventive measures to lower the company-wide risk index over time.")
        total_wb_count = len(incident)
        if incident:
            month = incident[0].get("month&Year", 'N/A')
        
        db_data = {
            "Total No of Incidents": {"Total Incidents": total_wb_count, "Month & Year": month},
            "Risk Index": {"Score": riskIndex, "Description": description},
            "No of Highly Exploitable Unique CVEs": count,
            "No of Incidents Closed without acknowledgement": total_closed_without_acknowledgement,
            "Top incidents": []
        }
        
        detailed_summary_data = [
            ["Total Incidents", "Month & Year", "", "Score", "Description", "", count, "", total_closed_without_acknowledgement, "", 
             "No of Incidents", "Incident Names with no of Occurrence", "Priority - Impact", "Data Source"]
        ]

        priority_map = {
            "critical": "P1 - Critical",
            "high": "P2 - High",
            "medium": "P3 - Medium",
            "low": "P4 - Low"
        }

        top_5_count = 0
        for incident in top_5_alert:
            incident_name_count = f"{incident['name']} - {incident['NoofOccurences']}"
            priority_level = priority_map.get(incident["severity"], "P4 - Low")
            data_source = incident["alertProvider"]
            
            detailed_summary_data.append(["", "", "", "", "", "", "", "", "", "", "", incident_name_count, priority_level, data_source])
            db_data["Top incidents"].append({"Incident Names with no of Occurrence": incident_name_count, "Priority - Impact": priority_level, "Data Source": data_source})
            top_5_count += incident["NoofOccurences"]

        db_data["Top incidents"].append({"No of Incidents": top_5_count})
        visionOne_data["Detailed Summary"] = db_data
        
        detailed_summary_sheet.merge_cells('A1:B1')
        detailed_summary_sheet['A1'] = "Total No of Incidents"

        detailed_summary_sheet.merge_cells('D1:E1')
        detailed_summary_sheet['D1'] = "Risk Index"

        detailed_summary_sheet['G1'] = "No of Highly Exploitable Unique CVEs"
        detailed_summary_sheet['I1'] = "No of Incidents Closed without acknowledgement"
        if len(detailed_summary_data) >= 2:
            detailed_summary_data[1][0] = total_wb_count
            detailed_summary_data[1][1] = month
            detailed_summary_data[1][3] = riskIndex
            detailed_summary_data[1][4] = description
        else:
            detailed_summary_data.append([total_wb_count, month, "", riskIndex, description])

        detailed_summary_sheet.merge_cells('K1:N1')
        detailed_summary_sheet['K1'] = "Top incidents"

        for row in detailed_summary_data:
            detailed_summary_sheet.append(row)

        detailed_summary_sheet.merge_cells('K3:K7')
        detailed_summary_sheet['K3'] = top_5_count

    except Exception as e:
        logging.error(f"Error Adding data to Detailed summary: {e}")
        return None
    
def xl_vulnerability_assessment_report():
    data = security_posture_data.get('Vulnerability Assesment', {})
    
    try:
        vulnerability_assessment_report_sheet = workbook.create_sheet(title="Vulnerability Assessment Report")

        count = data.get("count", "N/A")
        mttpDays = data.get("mttpDays", "N/A")
        averageUnpatchedDays = data.get("averageUnpatchedDays", "N/A")
        density = data.get("density", "")
        vulnerableEndpointRate = data.get("vulnerableEndpointRate", "N/A")
        legacyOsEndpointCount = data.get("legacyOsEndpointCount", "N/A")

        db_data = {
            "Internal Assets": {
                "Vulnerable Endpoint Percentage": vulnerableEndpointRate,
                "Highly-Exploitable CVE Density": density,
                "Devices with Legacy Windows Systems": legacyOsEndpointCount,
                "Average Unpatched Time": averageUnpatchedDays,
                "Mean Time to Patch (MTTP)": mttpDays,
                "Highly-Exploitable Unique CVEs": count
            }
        }
        visionOne_data["Vulnerability Assessment Report"] = db_data
        
        vulnerability_assessment_report_data = [
            ["Vulnerable Endpoint Percentage", vulnerableEndpointRate, "", "Highly Exploitable Unique CVES on Hosts"],
            ["Highly-Exploitable CVE Density", density, "", "Vulnerable Host Percentage"],
            ["Devices with Legacy Windows Systems", legacyOsEndpointCount, "", "Highly Exploitable CVE Density of Hosts"],
            ["Average Unpatched Time", averageUnpatchedDays],
            ["Mean Time to Patch (MTTP)", mttpDays],
            ["Highly-Exploitable Unique CVEs", count]
        ]

        vulnerability_assessment_report_sheet.merge_cells('A1:B1')
        vulnerability_assessment_report_sheet['A1'] = "Internal Assets"

        vulnerability_assessment_report_sheet.merge_cells('D1:E1')
        vulnerability_assessment_report_sheet['D1'] = "Internet-Facing Assets"

        vulnerability_assessment_report_sheet['G1'] = "Recommendation/Note/Summary"

        for row in vulnerability_assessment_report_data:
            vulnerability_assessment_report_sheet.append(row)
    except Exception as e:
        logging.error(f"Error Adding data to Vulnerability Report: {e}")
        return
    
def xl_system_configuration_report():
    try:
        data = security_posture_data.get('Security Configuration', {})
        insecure_host_count = security_posture_data.get("Insecure Host Count", "N/A")
        
        system_configuration_sheet = workbook.create_sheet(title="System Configuration Report")

        weakAuthenticationCount = data.get("weakAuthenticationCount", "N/A")
        increaseAttackSurfaceRiskCount = data.get("increaseAttackSurfaceRiskCount", "N/A")
        excessivePrivilegeCount = data.get("excessivePrivilegeCount", "N/A")
        
        db_Data = {
            "System Configuration Report": {
                "Accounts with Weak Authentication": weakAuthenticationCount,
                "Accounts that Increase Attack Surface Risk": increaseAttackSurfaceRiskCount,
                "Accounts with Excessive Privilege": excessivePrivilegeCount,
                "Hosts with Insecure Connection Issues": insecure_host_count
            }
        }
        visionOne_data["System Configuration Report"] = db_Data

        system_configuration_data = [
            ["Accounts with Weak Authentication", weakAuthenticationCount],
            ["Accounts that Increase Attack Surface Risk", increaseAttackSurfaceRiskCount],
            ["Accounts with Excessive Privilege", excessivePrivilegeCount],
            ["Legacy Authentication Protocol with Log On Activity"],
            ["Unexpected Internet-Facing Services/Ports"],
            ["Hosts with Insecure Connection Issues", insecure_host_count]
        ]

        # Merge and setup cell values
        system_configuration_sheet.merge_cells('A1:B1')
        system_configuration_sheet['A1'] = "System Configuration Report"
        system_configuration_sheet['D1'] = "Recommendation/Note/Summary"

        # Append rows to the sheet
        for row in system_configuration_data:
            system_configuration_sheet.append(row)
            
    except Exception as e:
        logging.error(f"Error Adding data to System Configuration Report: {e}")
        return None
    
def xl_agent_version_summary():
    try:
        data = security_posture_data.get('Agent Life Cycle', {})
        agent_version_summary_sheet = workbook.create_sheet(title="Agent Version Summary")

        latestCount = data.get("latestCount", "N/A")
        outdatedCount = data.get("outdatedCount", "N/A")
        otherCount = data.get("otherCount", "N/A")
        if str(latestCount).isdigit() and str(otherCount).isdigit() and str(outdatedCount).isdigit():
            total = latestCount + outdatedCount + otherCount
        else:
            total = "N/A"
        db_Data = {
            "All Agents Version Summary": {
                "Total (Endpoint + Server)": total,
                "Latest Version (Endpoint + Server)": latestCount,
                "Older Version (Endpoint + Server)": otherCount,
                "End-of-Life version (Endpoint + Server)": outdatedCount
            }
        }
        visionOne_data["Agent Version Summary"] = db_Data
        
        agent_version_summary_data = [
            ["Total (Endpoint + Server)", total, "", "Total Servers", "", "", "Total Endpoints"],
            ["Latest Version (Endpoint + Server)", latestCount, "", "Latest Version Servers", "", "", "Latest Version Endpoints"],
            ["Older Version (Endpoint + Server)", otherCount, "", "Older Version Servers", "", "", "Older Version Endpoints"],
            ["End-of-Life version (Endpoint + Server)", outdatedCount, "", "End-of-Life version Servers", "", "", "End-of-Life version Endpoints"]
        ]

        agent_version_summary_sheet.merge_cells('A1:B1')
        agent_version_summary_sheet['A1'] = "All Agents Version Summary"

        agent_version_summary_sheet.merge_cells('D1:E1')
        agent_version_summary_sheet['D1'] = "Server & Workload Protection"

        agent_version_summary_sheet.merge_cells('G1:H1')
        agent_version_summary_sheet['G1'] = "Standard Endpoint Protection"

        agent_version_summary_sheet['J1'] = "Recommendation/Note/Summary"

        for row in agent_version_summary_data:
            agent_version_summary_sheet.append(row)
    except Exception as e:
        logging.error(f"Error Adding data to Agent Version Summary: {e}")
        return
         
def xl_detection_summary_from_a1():
    malware_data = get_syslog_data("officescan_virus", ["File cleaned", "File quarantined", "File deleted", "Restart action required"])
    spyware_data = get_syslog_data("spyware", ["File cleaned"])
    cnc_data = get_syslog_data("ncie", ["Block"])
    intrusion_data = get_syslog_data("intrusion_prevention", ["Block"])
    try:
        file_cleaned_malware = malware_data["total_count"][0]["File cleaned"]
        file_cleaned_spyware = spyware_data["total_count"][0]["File cleaned"]
        total_cleaned = file_cleaned_malware + file_cleaned_spyware
        
        file_quarantine_malware = malware_data["total_count"][1]["File quarantined"]
        file_deleted_malware = malware_data["total_count"][2]["File deleted"]
        file_restarted_malware = malware_data["total_count"][3]["Restart action required"]
        intrusion_count = intrusion_data["total_count"][0]["Block"]
        cnc_count = cnc_data["total_count"][0]["Block"]
        
        malware_clean_top_3 = malware_data["top_3_endpoints"][0].get("File cleaned", [])
        malware_quarantine_top_3 = malware_data["top_3_endpoints"][1].get("File quarantined", [])
        malware_deleted_top_3 = malware_data["top_3_endpoints"][2].get("File deleted", [])
        spyware_clean_top_3 = spyware_data["top_3_endpoints"][0].get("File cleaned", [])
        cnc_block_top_3 = cnc_data["top_3_endpoints"][0].get("Block", [])
        intrusion_block_top_3 = intrusion_data["top_3_endpoints"][0].get("Block", [])
            
        db_data = {
            "Detection Summary form Apex One": {
                "Virus/Malware & Spyware/Grayware": {
                    "File Cleaned (Virus/Malware + Spyware/Grayware)": total_cleaned,
                    "File Quarantine (Virus / Malware)": file_quarantine_malware,
                    "Restart Action Required (Virus / Malware)": file_restarted_malware,
                    "File Deleted (Virus / Malware)": file_deleted_malware
                },
                "C & C Connections & Intrusion attempts Blocked": {
                    "Total C & C Connections Blocked": cnc_count,
                    "Total Intrusion Attempts Blocked": intrusion_count
                }
            },
            "Top 03 Endpoints": {
                "File Cleaned/Spyware": [],
                "File Cleaned/Malware": [],
                "File Qurantined(Malware)": [],
                "File Deleted": [],
                "C & C Connection Blocked": [],
                "Intrusion Attempts Blocked": []
            }
        }
        
        visionOne_data["Detection Summary from A1"] = db_data
        
        detection_summary_sheet = workbook.create_sheet(title="Detection Summary from A1")
        
        detection_summary_data = [
            ["", "", "", "", "", "File Cleaned(Malware)", "File Cleaned(Spyware)", "File Quarantined", "File Deleted", "C&C Connection Blocked", "Intrusion Attempts Blocked"],
            ["File Cleaned (Virus/Malware + Spyware/Grayware)", total_cleaned, "Total C & C connections Blocked", cnc_count, ""],
            ["File Quarantine (Virus / Malware)", file_quarantine_malware, "Total Intrusion attempts Blocked", intrusion_count, ""],
            ["Restart Action Required (Virus / Malware)", file_restarted_malware, "", "", ""],
            ["File Deleted (Virus / Malware)", file_deleted_malware, "", "", ""]
        ]
        
        for i in range(3):            
            malware_clean_top_3_value = malware_clean_top_3[i][0] + " - " + str(malware_clean_top_3[i][1]) if i < len(malware_clean_top_3) else "NA"
            malware_quarantine_top_3_value = malware_quarantine_top_3[i][0] + " - " + str(malware_quarantine_top_3[i][1]) if i < len(malware_quarantine_top_3) else "NA"
            malware_deleted_top_3_value = malware_deleted_top_3[i][0] + " - " + str(malware_deleted_top_3[i][1]) if i < len(malware_deleted_top_3) else "NA"
            spyware_clean_top_3_value = spyware_clean_top_3[i][0] + " - " + str(spyware_clean_top_3[i][1]) if i < len(spyware_clean_top_3) else "NA"
            cnc_block_top_3_value = cnc_block_top_3[i][0] + " - " + str(cnc_block_top_3[i][1]) if i < len(cnc_block_top_3) else "NA"
            intrusion_block_top_3_value = intrusion_block_top_3[i][0] + " - " + str(intrusion_block_top_3[i][1]) if i < len(intrusion_block_top_3) else "NA"
            
            db_data["Top 03 Endpoints"]["File Cleaned/Spyware"].append(spyware_clean_top_3_value)
            db_data["Top 03 Endpoints"]["File Cleaned/Malware"].append(malware_clean_top_3_value)
            db_data["Top 03 Endpoints"]["File Deleted"].append(malware_deleted_top_3_value)
            db_data["Top 03 Endpoints"]["C & C Connection Blocked"].append(cnc_block_top_3_value)
            db_data["Top 03 Endpoints"]["Intrusion Attempts Blocked"].append(intrusion_block_top_3_value)
            db_data["Top 03 Endpoints"]["File Qurantined(Malware)"].append(malware_quarantine_top_3_value)
            
            detection_summary_data[i+1].extend([
                malware_clean_top_3_value, spyware_clean_top_3_value,
                malware_quarantine_top_3_value, malware_deleted_top_3_value,
                cnc_block_top_3_value, 
                intrusion_block_top_3_value
            ]) 
            
        detection_summary_sheet.merge_cells('A1:D1')
        detection_summary_sheet['A1'] = "Detection Summary form Apex One"
        
        detection_summary_sheet.merge_cells('F1:J1')
        detection_summary_sheet['F1'] = "Top 03 Endpoints"
        
        detection_summary_sheet['L1'] = "Recommendation/Note/Summary"

        for row in detection_summary_data:
            detection_summary_sheet.append(row)
            
        detection_summary_sheet.merge_cells('A2:B2')
        detection_summary_sheet['A2'] = "Virus/Malware & Spyware/Grayware"
        
        detection_summary_sheet.merge_cells('C2:D2')
        detection_summary_sheet['C2'] = "C & C Connections & Intrusion attempts Blocked"
        
    except Exception as e:
        logging.error(f"Error Adding data to Detection Summary: {e}")
        return
    
def xl_slo_summary():
    data = get_soar_alerts("sla")
    try:
        slo_summary_sheet = workbook.create_sheet(title="SLO Summary")
        
        
        slo_summary_data = [
            ["Total No of Incidents Closed", "SLO Met", "SLO not Met"]
        ]
        
        slo_met_count = 0
        slo_not_met_count = 0
        for alerts in data:
            for alert in alerts["query_result"]:
                if "Closed" in alert["grouping"]:
                    slo_met_count += alert["SLA_Met"]
                    slo_not_met_count += alert["SLA_Not_Met"]
        
        total_count = slo_met_count + slo_not_met_count
        slo_summary_data.append([total_count, slo_met_count, slo_not_met_count])
        
        db_data = {"SLO Summary": {"Total No of Incidents Closed": total_count, "SLO Met": slo_met_count, "SLO Not Met": slo_not_met_count}}
        visionOne_data["SLO Summary"] = db_data
         
        slo_summary_sheet.merge_cells('A1:C1')
        slo_summary_sheet['A1'] = "SLO Summary"

        for row in slo_summary_data:
            slo_summary_sheet.append(row)
    except Exception as e:
        logging.error(f"Error Adding data to SLO Summary: {e}")
        return

def xl_V1_workbench_incidents_summary():
    data = get_incident_summary_alerts("sla-by-v1-priority", "priority", ['P1', 'P2', 'P3', 'P4'])
    try:
        V1_workbench_incidents_summary_sheet = workbook.create_sheet(title="V1 Workbench Incidents Summary")
        
        V1_workbench_incidents_summary_data = [
            ["Priorities", "Incidents Closed with Resolution", "Incidents Closed without Acknowledgement", "Pending Incidents with Customer", "Pending Incidents with SOC Team"]
        ]
        
        db_data = {
            "V1 Workbench Incidents Summary": {
                "P1": {},
                "P2": {},
                "P3": {},
                "P4": {}
            }
        }
        priorities = ["P1", "P2", "P3", "P4"]
        
        for priority in priorities:
            closed_with_resolution = data["closed_with_resolution"][priority]
            closed_without_acknowledgement = data["closed_without_acknowledgement"][priority]
            pending_with_soc_team = data["pending_with_soc_team"][priority]
            V1_workbench_incidents_summary_data.append([priority, closed_with_resolution, closed_without_acknowledgement, "", pending_with_soc_team])
            db_data["V1 Workbench Incidents Summary"][priority].update({"Incidents Closed with Resolution": closed_with_resolution, "Incidents Closed without Acknowledgement": closed_without_acknowledgement, "Pending Incidents with SOC Team": pending_with_soc_team})
        
        visionOne_data["V1 Workbench Incidents Summary"] = db_data    
               
        V1_workbench_incidents_summary_sheet.merge_cells('A1:E1')
        V1_workbench_incidents_summary_sheet['A1'] = "V1 Workbench Incidents Summary"

        for row in V1_workbench_incidents_summary_data:
            V1_workbench_incidents_summary_sheet.append(row)
    except Exception as e:
        logging.error(f"Error Adding data to V1 Workbench Incidents Summary: {e}")
        return

def xl_siem_incidents_summary():
    data = get_incident_summary_alerts("sla-by-third-party-priority", "priority", ['P1', 'P2', 'P3', 'P4'])
    try:
        siem_incidents_summary_sheet = workbook.create_sheet(title="SIEM Incidents Summary")
        
        siem_incidents_summary_data = [
            ["Priorities", "Incidents Closed with Resolution", "Incidents Closed without Acknowledgement", "Pending Incidents with Customer", "Pending Incidents with SOC Team"]
        ]
        
        db_data = {
            "SIEM Incidents Summary": {
                "P1": {},
                "P2": {},
                "P3": {},
                "P4": {}
            }
        }
        priorities = ["P1", "P2", "P3", "P4"]
        
        for priority in priorities:
            closed_with_resolution = data["closed_with_resolution"][priority]
            closed_without_acknowledgement = data["closed_without_acknowledgement"][priority]
            pending_with_soc_team = data["pending_with_soc_team"][priority]
            siem_incidents_summary_data.append([priority, closed_with_resolution, closed_without_acknowledgement, "", pending_with_soc_team])
            db_data["SIEM Incidents Summary"][priority].update({"Incidents Closed with Resolution": closed_with_resolution, "Incidents Closed without Acknowledgement": closed_without_acknowledgement, "Pending Incidents with SOC Team": pending_with_soc_team})
        
        visionOne_data["SIEM Incidents Summary"] = db_data 
                   
        siem_incidents_summary_sheet.merge_cells('A1:E1')
        siem_incidents_summary_sheet['A1'] = "SIEM Incidents Summary"

        for row in siem_incidents_summary_data:
            siem_incidents_summary_sheet.append(row)
    except Exception as e:
        logging.error(f"Error Adding data to SIEM Incidents Summary: {e}")
        return

def xl_overall_incidents_summary():
    data_third = get_incident_summary_alerts("sla-by-third-party-priority", "priority", ['P1', 'P2', 'P3', 'P4'])
    data_v1 = get_incident_summary_alerts("sla-by-v1-priority", "priority", ['P1', 'P2', 'P3', 'P4'])

    try:
        overall_incidents_summary_sheet = workbook.create_sheet(title="Overall Incidents Summary")
        
        db_data = {
            "Status Summary": {},
            "Overall Incidents Summary": {
                "P1": {},
                "P2": {},
                "P3": {},
                "P4": {}
            }
        }
        
        overall_incidents_summary_data = [
            ["Incidents Closed with Resolution", "", "", "Priorities", "Incidents Closed with Resolution", "Incidents Closed without Acknowledgement", "Pending Incidents with Customer", "Pending Incidents with SOC Team"]
        ]
        
        total_closed_with_resolution = 0
        global total_closed_without_acknowledgement
        total_closed_without_acknowledgement = 0
        total_pending_with_soc_team = 0
        
        priorities = ["P1", "P2", "P3", "P4"]
        for priority in priorities:
            closed_with_resolution = data_third["closed_with_resolution"][priority] + data_v1["closed_with_resolution"][priority]
            total_closed_with_resolution += closed_with_resolution
            
            closed_without_acknowledgement = data_third["closed_without_acknowledgement"][priority] + data_v1["closed_without_acknowledgement"][priority]
            total_closed_without_acknowledgement += closed_without_acknowledgement
            
            pending_with_soc_team = data_third["pending_with_soc_team"][priority] + data_v1["pending_with_soc_team"][priority]
            total_pending_with_soc_team += pending_with_soc_team
            
            overall_incidents_summary_data.append(["", "", "", priority, closed_with_resolution, closed_without_acknowledgement, "", pending_with_soc_team])
            db_data["Overall Incidents Summary"][priority].update({"Incidents Closed with Resolution": closed_with_resolution, "Incidents Closed without Acknowledgement": closed_without_acknowledgement, "Pending Incidents with SOC Team": pending_with_soc_team})
        
        db_data["Status Summary"] = {"Incidents Closed with Resolution": total_closed_with_resolution, "Incidents Closed without Acknowledgement": total_closed_without_acknowledgement, "Pending Incidents with SOC Team": total_pending_with_soc_team}
        visionOne_data["Overall Incidents Summary"] = db_data
                   
        overall_incidents_summary_sheet.merge_cells('D1:H1')
        overall_incidents_summary_sheet['D1'] = "Overall Incidents Summary"

        overall_incidents_summary_sheet.merge_cells('A1:B1')
        overall_incidents_summary_sheet['A1'] = "Status Summary"
        
        for row in overall_incidents_summary_data:
            overall_incidents_summary_sheet.append(row)
            
        overall_incidents_summary_sheet["B2"] = total_closed_with_resolution
        overall_incidents_summary_sheet["B3"] = total_closed_without_acknowledgement
        overall_incidents_summary_sheet["B5"] = total_pending_with_soc_team
        overall_incidents_summary_sheet["A3"] = "Incidents Closed without Acknowledgement"
        overall_incidents_summary_sheet["A4"] = "Pending Incidents with Customer"
        overall_incidents_summary_sheet["A5"] = "Pending Incidents with SOC Team"
        
    except Exception as e:
        logging.error(f"Error Adding data to Overall Incidents Summary: {e}")
        return

def xl_pending_incidents_summary():
    data = pending_incidents()
    try:
        pending_incidents_summary_sheet = workbook.create_sheet(title="Pending Incidents Summary")

        db_data = {"Pending Incidents Summary": []}
        
        pending_incidents_summary_sheet.merge_cells('A1:C1')
        pending_incidents_summary_sheet['A1'] = "Pending Incidents Summary"
        
        pending_incidents_summary_sheet.append(["Incident Names", "Priority", "No of Occurrence"])

        for incidents in data:
            pending_incidents_summary_sheet.append([incidents["title"], incidents["priority"], incidents["count"]])
            db_data["Pending Incidents Summary"].append({"Incident Name": incidents["title"], "Priority": incidents["priority"], "No of Occurance": incidents["count"]})
        
        visionOne_data["Pending Incidents Summary"] = db_data              
        
    except Exception as e:
        logging.error(f"Error Adding data to Pending Incidents Summary: {e}")

def xl_endpoint_feature_complicance():
    endpoint_details = [
        ("XDR Endpoint Sensor Not Enabled", "edrSensorStatus eq disabled"),
        ("Vulnerability Protection not Enabled", "eppAgentVulnerabilityProtection eq disabled"),
        ("XDR Endpoint Sensor, Vulnerability Protection Enabled", "edrSensorStatus eq enabled or eppAgentVulnerabilityProtection eq enabled")
    ]

    endpoint_data = []
    db_data = {"Endpoint Feature Compliance": {}}
    for label, query in endpoint_details:
        count = get_endpoint_details(query)
        endpoint_data.append([label, count])
        db_data["Endpoint Feature Compliance"].update({label: count})

    visionOne_data["Endpoint Feature Compliance"] = db_data
    try:
        endpoint_sheet = workbook.create_sheet(title="Endpoint Feature Complicance")
        
        endpoint_sheet.merge_cells('A1:B1')
        endpoint_sheet['A1'] = "Endpoint Feature Complicance"
        endpoint_sheet['D1'] = "Recommendation/Note/Summary"

        for row in endpoint_data:
            endpoint_sheet.append(row)
    except Exception as e:
        logging.error(f"Error Adding data to Endpoint Feature Complicance: {e}")

def xl_email_summary():
    data = get_quarantine_events()
    top3_data = get_sweeping_mails()

    senders = top3_data['top3_senders']
    receipts = top3_data['top3_receipts']
    try:
        db_data = {
            "Email Status": data['email_status'],
            "Top 03 Sender and Receipts": {"Top 3 Senders": senders, "Top 3 Receipts": receipts},
            "Threat Type": data['threat_type']
        }
        visionOne_data["Email Summary"] = db_data
        
        email_data = [
            ['Quarantined', 'Delete Success', '', 'Top 03 Sender', 'Top 03 Receipts', '', 'Malicious Files', data['threat_type']['Malicious Files']],
            [data['email_status']['Quarantined'], data['email_status']['Deleted'], '', senders[0] if len(senders) > 0 else '', receipts[0] if len(receipts) > 0 else '', '', 'Malicious URLs', data['threat_type']['Malicious URLs']],
            ['','','', senders[1] if len(senders) > 1 else '', receipts[1] if len(receipts) > 1 else '','','Phishing', data['threat_type']['Phishing']],
            ['','','', senders[2] if len(senders) > 2 else '', receipts[2] if len(receipts) > 2 else '','', 'Spoofing', data['threat_type']['Spoofing']],
            ['','','','','','', 'Suspicious Object', data['threat_type']['Suspicious Object']],
            ['','','','','','', 'Blocked Object', data['threat_type']['Blocked Object']]
        ]
        email_sheet = workbook.create_sheet(title="Email Summary")
        
        email_sheet.merge_cells('A1:B1')
        email_sheet['A1'] = "Email Status"
        email_sheet.merge_cells('D1:E1')
        email_sheet['D1'] = "Top 03 Sender and Receipts"
        email_sheet.merge_cells('G1:H1')
        email_sheet['G1'] = "Threat Type"
        email_sheet['J1'] = "Recommendation/Note/Summary"

        for row in email_data:
            email_sheet.append(row)
    except Exception as e:
        logging.error(f"Error Adding data to Email Summary: {e}")

def xl_risk_matrics():
    current_month_risk_Score = security_posture_data.get('Risk Index')["riskIndex"]
    previous_month_names = get_previous_three_months(month_name)
    previous_risk_scores = []
    try:
        for i in range(2):
            if connection_check:
                is_file_present, file_id, document = get_id_of_es_file(f"{tenant_name}_{previous_month_names[i]}_Report")
                if is_file_present:
                    previous_risk_scores.append(document.get('document', {}).get('Detailed Summary', {}).get('Risk Index', {}).get('Score'))
                else:
                    previous_risk_scores.append('NA')
            else:
                previous_risk_scores.append('NA')
    
        db_data = {
            "Last Three Month Risk Score": {
                previous_month_names[0]: previous_risk_scores[0],
                previous_month_names[1]: previous_risk_scores[1],
                month_name: current_month_risk_Score
                }
        }
        visionOne_data["Risk Matrics"] = db_data
        previous_risk_scores.append(current_month_risk_Score)
        risk_matrics_data = [
            ["", "", "", "", "Files/Hash", "URL", "Email", "IP"],
            previous_month_names, previous_risk_scores
        ]
        
        risk_matrics_sheet = workbook.create_sheet(title="Risk Matrics")
        
        risk_matrics_sheet.merge_cells('A1:C1')
        risk_matrics_sheet['A1'] = "Last Three Month Risk Score"
        risk_matrics_sheet.merge_cells('E1:H1')
        risk_matrics_sheet['E1'] = "Top Risk Vectors based on Incidents detected"
        
        for row in risk_matrics_data:
            risk_matrics_sheet.append(row)
            
        risk_matrics_sheet.merge_cells('A2:C2')
        risk_matrics_sheet['A2'] = "Name of Previous Months"
    except Exception as e:
        logging.error(f"Error Adding data to Risk Matrics: {e}")

def xl_top_vulnerability_detected():
    data = get_top10_cves()
    try:
        top_vulnerability_detected_sheet = workbook.create_sheet(title="Top Vulnerability Detected")

        db_data = {"Top Vulnerabilities Detected": []}
        
        top_vulnerability_detected_sheet.merge_cells('A1:C1')
        top_vulnerability_detected_sheet['A1'] = "Top Vulnerabilities Detected"

        top_vulnerability_detected_sheet.append(["CVEs", "CVE impact score", "Impact scope"])

        for cve in data:
            top_vulnerability_detected_sheet.append([cve["cve_id"], cve["score"]])
            db_data["Top Vulnerabilities Detected"].append({"CVE": cve["cve_id"], "CVE impact score": cve["score"]})
        
        visionOne_data["Top Vulnerability Detected"] = db_data
          
    except Exception as e:
        logging.error(f"Error Adding data to Top Vulnerability Detected: {e}")

def xl_product_assessment_report():
    get_v1_status = isvalid_token().get('status', '')
    get_c1_status = make_cloud_one_api_call("APIKeysApi", "describe_current_api_key").get("_active", '')
    v1_status = "Connected" if get_v1_status == "available" else "Not Connected"
    c1_status = "Connected" if get_c1_status == True else "Not Connected"
    try:
        product_assessment_report_sheet = workbook.create_sheet(title="Product Assessment Report")

        product_assessment_report_sheet['A1'] = "TM Products Summary"
        product_assessment_report_sheet['E1'] = "Third Party Products Summary"

        header = ["TM Product", "Connection Status", "Identifier", "", "Product", "Connection Status", "Remarks"]
        product_assessment_report_sheet.append(header)

        product_assessment_report_data = [
            ["Vision One", v1_status],
            ["Cloud One", c1_status],
            ["Apex Central"],
            ["Cloud App security"]
        ]
        
        for row in product_assessment_report_data:
            product_assessment_report_sheet.append(row)  
            
        db_data = {
            "TM Products Summary": [
                {"TM Product": "Vision One", "Connection Status": v1_status},
                {"TM Product": "Cloud One", "Connection Status": c1_status},
                {"TM Product": "Apex Central", "Connection Status": ""},
                {"TM Product": "Cloud App security", "Connection Status": ""}
            ],
            "Third Party Products Summary": []
        }  
        
        visionOne_data["Product Assessment Report"] = db_data
                
        product_assessment_report_sheet.merge_cells('A1:C1')
        product_assessment_report_sheet.merge_cells('E1:G1')
        
    except Exception as e:
        logging.error(f"Error Adding data to Pending Incidents Summary: {e}")
        
######################################### Weekly Excel Format Functions #######################################

def xl_weekly_executive_summary(start_date, end_date):
    total_incidents = incident_overview()
    top_5_alert = top_5_alerts()
    endpoint_sensor= get_endpoint_details("edrSensorStatus eq enabled")
    endpoint_protection = get_endpoint_details("eppAgentVulnerabilityProtection eq enabled")
    total_endpoint_sensor = get_endpoint_details("edrSensorStatus eq enabled or edrSensorStatus eq disabled")
    total_endpoint_protection = get_endpoint_details("eppAgentVulnerabilityProtection eq enabled or eppAgentVulnerabilityProtection eq disabled")
    try:
        top5_endpoints = get_top5_endpoints()
    except Exception as e:
        logging.error(f"Error Getting Top 5 Endpoints: {e}")
        top5_endpoints = []
    
    try:
        executive_summary_sheet = workbook.create_sheet(title="Executive Summary")

        executive_summary_data = [
            ["total_incidents", "", "", "", "risk_index", "", "highly_exploitable", "", "Incident_Closed", "", "Highest_incidient", "", "Highest_incidient_date", "", "top_incident", "", "", "", "agent_life_cycle", "", "", "","", "top_endpoint", "", "", "", "", "endpoint_protection", "endpoint_sensor", "", "R_S_N"],
            ["total_incidents", "start_date", "end_date", "", "risk_persentage", "", "no_highly_exploitable", "", "Incident_Closed", "", "Highest_incidient", "", "Highest_incidient_date", "", "number_of_incidents", "incidents_names", "data_source_from_v1", "", "Standard_Endpoint_Protection", "", "Server&workload_security", "", "", "number_of_endpoint", "endpoints_names", "number_of_detection_with_severity", "action_taken_from_soc_team", "", "value", "value", "", "summary"]
        ]

        count = len(total_incidents)
        
        end_date = datetime.strptime(end_date, '%Y-%m-%dT%H:%M:%SZ').strftime('%d/%m/%y')
        start_date = datetime.strptime(start_date, '%Y-%m-%dT%H:%M:%SZ').strftime('%d/%m/%y')
        if top_5_alert:
            highest_incident_count = top_5_alert[0].get("NoofOccurences", 'NA') if top_5_alert else 'NA'
            highest_incident_date = top_5_alert[0].get('createdDateTime', '') if top_5_alert else ''
            if highest_incident_date:
                highest_incident_date = datetime.strptime(highest_incident_date, '%Y-%m-%dT%H:%M:%SZ').strftime('%d/%m/%y')
        else:
            highest_incident_count = 'N/A'
            highest_incident_date = 'N/A'
        risk_index = security_posture_data.get("Risk Index", {}).get("riskIndex", 'N/A')
        high_ex_cves = security_posture_data.get("Highly Exploitable Unique CVES", 'N/A')            
        
        db_data = {
            "Total Incidents": {"Total Incidents": count, "Start_date": start_date, "End_date": end_date, },
            "Risk Index": {"Risk Percentage": risk_index},
            "Highly Exploitable no of CVE's": high_ex_cves,
            "Highest_incident": {"highest_incident": highest_incident_count, "highest incident date": highest_incident_date},
            "Top incident": [],
            "Endpoint Details": {"endpoint_protection_enabled": endpoint_protection, "endpoint_sensor_enabled": endpoint_sensor, "total_endpoint_protection_enabled": total_endpoint_protection, "total_endpoint_sensor_enabled": total_endpoint_sensor},
            "Total Closed Incidents": total_closed_incidents,
            "Top Endpoints": []
        }

        top_5_count = 0
        if top_5_alert:
            for incident in top_5_alert:
                incident_name_count = f"{incident['name']} - {incident['NoofOccurences']}"
                data_source = incident["alertProvider"]
                executive_summary_data.append(["", "", "", "", "", "", "", "", "", "", "","","","","", incident_name_count, data_source])
                db_data["Top incident"].append({"Incident Name": incident_name_count, "Data Source": data_source})
                top_5_count += incident["NoofOccurences"]
        else:
            for i in range(5):
                executive_summary_data[i+2].append(["", "", "", "", "", "", "", "", "", "", "","","","","", "", ""])
            
        executive_summary_data[2].extend(["", "label", "value", "label", "value", ""])
        executive_summary_data[3].extend(["", "Latest Version", "", "Latest Version", "", ""])
        executive_summary_data[4].extend(["", "Older Version", "", "Older Version", "", ""])
        executive_summary_data[5].extend(["", "End of Life", "", "End of Life", "", ""])
        executive_summary_data[6].extend(["", "", "", "", "", ""])
        
        if top5_endpoints:
            c = 2
            for endpoint in top5_endpoints:
                no_detection_severity = ''
                risklevel = endpoint.get("riskLevels", "")
                if risklevel:
                    for key, val in risklevel.items():
                        no_detection_severity += str(val) + " " + key + " "
                executive_summary_data[c].extend(["", endpoint.get("endpoinName", ""), no_detection_severity[:-1]])
                db_data["Top Endpoints"].append({"Endpoint Name": endpoint.get("endpoinName", ""), "No of Detection with Severity": no_detection_severity[:-1]})
                c += 1
            executive_summary_data[2][23] = len(top5_endpoints)
        else:
            executive_summary_data[2].extend(["", "", ""])
            executive_summary_data[3].extend(["", "", ""])
            
        executive_summary_data[2].extend(["", "", endpoint_protection, endpoint_sensor])
        executive_summary_data[3].extend(["", "", total_endpoint_protection, total_endpoint_sensor])
        
        visionOne_data["Executive Summary"] = db_data
        
        executive_summary_data[2][0] = count
        executive_summary_data[2][1] = start_date
        executive_summary_data[2][2] = end_date
        executive_summary_data[2][4] = risk_index
        executive_summary_data[2][6] = high_ex_cves
        executive_summary_data[2][8] = total_closed_incidents
        executive_summary_data[2][10] = highest_incident_count
        executive_summary_data[2][12] = highest_incident_date
        
        for row in executive_summary_data:
            executive_summary_sheet.append(row)
        
        executive_summary_sheet.merge_cells('O3:O7')
        executive_summary_sheet['O3'] = top_5_count     
    except Exception as e:
        logging.error(f"Error Adding data to Executive summary: {e}")
        return 

def xl_weekly_slo_summary():
    data = get_soar_alerts("sla")
    try:
        slo_summary_sheet = workbook.create_sheet(title="SLO Summary")        
        
        slo_met_count = 0
        slo_not_met_count = 0
        for alerts in data:
            for alert in alerts["query_result"]:
                if "Closed" in alert["grouping"]:
                    slo_met_count += alert["SLA_Met"]
                    slo_not_met_count += alert["SLA_Not_Met"]
        
        total_count = slo_met_count + slo_not_met_count
        slo_summary_data = [
            ["Graph", "", "", "", "SLO Details"], ["Key", "Case Volume", "", "", "Priority", "Description", "Response Time"],
            ["Total Closed Incidents", total_count ],
            ["SLO Met", slo_met_count],
            ["SLO not Met", slo_not_met_count]
        ]
        
        db_data = {"SLO Summary": {"Total No of Incidents Closed": total_count, "SLO Met": slo_met_count, "SLO Not Met": slo_not_met_count}}
        visionOne_data["SLO Summary"] = db_data

        for row in slo_summary_data:
            slo_summary_sheet.append(row)
    except Exception as e:
        logging.error(f"Error Adding data to SLO Summary: {e}")
        return

def xl_weekly_pending_incidents_summary():
    data = pending_incidents()
    try:
        pending_incidents_summary_sheet = workbook.create_sheet(title="Pending Incidents Summary")
        pending_incidents_summary_data = [
            ["Total Pending Incidents", "Pending Incidents from Customer", "Pending Incidents from SOC Team", "", "Piechart", "", "", "R_S_N"],
            [len(data), "NA", len(data),"", "Key", "Series 1", "", "Recommendation"],
            ["", "", "", "","Total Pending Incidents", len(data)],
            ["", "", "", "", "Pending Incidents from Customer", "NA"],
            ["", "", "", "", "Pending Incidents from SOC Team", len(data)]
        ]
        
        for row in pending_incidents_summary_data:
            pending_incidents_summary_sheet.append(row)
        
        db_data = {"Total Pending Incidents": len(data), "Pending Incidents from Customer": "NA", "Pending Incidents from SOC Team": len(data)}
        visionOne_data["Pending Incidents Summary"] = db_data
                

    except Exception as e:
        print(f"Error Adding data to Pending Incidents Summary: {e}")

def xl_weekly_endpoint_inventory():
    action_required = get_endpoint_details("availableActions eq immediateActionRequired")
    get_v1_status = isvalid_token().get('status', '')
    get_c1_status = make_cloud_one_api_call("APIKeysApi", "describe_current_api_key").get("_active", '')
    v1_status = "Connected" if get_v1_status == "available" else "Not Connected"
    c1_status = "Connected" if get_c1_status == True else "Not Connected"
    try:
        endpoint_inventory_sheet = workbook.create_sheet(title="Endpoint Inventory")
       
        endpoint_inventory_data = [
            ["Bar_graph", "",  "", "R_S_N", "", "Connected_Products_and_License_Information", "" ],
            ["key", "info",  "", "recommendation", "", "License Information", "", "Products Connected", ""],
            ["XDR feature enabled", "",  "", "", "", "Status", "Products", "Status", "Product"],
            ["XDR feature not enabled","", "", "", "", "", "Vision One", v1_status, "Vision One"],
            ["Action Required", action_required,"", "", "", "", "Cloud One", c1_status, "Cloud One"],
            ["", "","", "", "", "", "Apex Central", "", "Apex Central"],
            ["", "","", "", "", "", "Cloud App Security", "", "Cloud App Security"]
        ]
       
        for row in endpoint_inventory_data:
            endpoint_inventory_sheet.append(row)
        
       
        db_data = {
            "Bar_graph": {"XDR feature enabled": "", "XDR feature not enabled": "", "Action Required": action_required},
            "Connected_Products_and_License_Information":{
                "Products Connected":[
                    {"Status": v1_status, "Product":"Vision One"},
                    {"Status": c1_status, "Product":"Cloud One"},
                    {"Status": "", "Product":"Apex Central"},
                    {"Status": "", "Product":"Cloud App Security"}
                ]
            }
        }
       
        visionOne_data["Endpoint inventory"] = db_data
       
    except Exception as e:
        logging.error(f"Error Adding data to Endpoint Inventory: {e}")
        return

def xl_weekly_threat_intel_summary():
    data_third = get_incident_summary_alerts("sla-by-third-party-priority", "priority", ["P1", "P2", "P3", "P4"])
    data_v1 = get_incident_summary_alerts("sla-by-v1-priority", "priority", ["P1", "P2", "P3", "P4"])
    categories = get_incident_categories(get_soar_alerts("sla-by-category"))
    data_by_category = get_incident_summary_alerts("sla-by-category", "category", categories)
    data_by_severity = get_soar_alerts("cases-by-impact")[0].get('query_result', [])
    try:
        threat_intel_summary_sheet = workbook.create_sheet(title="Threat Intel Summary")
       
        db_data = {
            "Incident Summary by status": {},
            "Incident Summary by priority": {
                "P1": {},
                "P2": {},
                "P3": {},
                "P4": {}
            },
            "Incident Summary by category":{},
            "Incident Summary by severity": data_by_severity
        }
       
        threat_intel_summary_data = [
            ["Key", "IOC Sweeped", "IOC Matched", "", "recommendation", "", "Sr.No", "Advisory Name", "Matched IOC type", "Matched IOC Details", "No of Endpoints / Email", "", "Key", "Series 1", "", "notes", "", "Key", "Series 1", "", "Summary", "", "Key", "Closed", "Pending from SOC", "Pending from Customer", "", "recommendation", "", "Key", "Closed", "Pending from SOC", "Pending from Customer", "", "recommendation"],
            ['IP', "", "","","","","","","","","","","Critical", "NA", "","","","Pending Incidents","","","",""],
            ['URL', "", "","","","","","","","","","","High", "NA", "","","","Closed Incidents","","","",""],
            ['Hash', "", "","","","","","","","","","","Medium", 'NA', "","","","","","","",""],
            ['Domain', "", "","","","","","","","","","","low", 'NA', "","","","","","","",""]
        ]
        for i in data_by_severity:
            if i['impact'] == 'Low':
                threat_intel_summary_data[4][13] = i['count']
            elif i['impact'] == 'Medium':
                threat_intel_summary_data[3][13] = i['count']
            elif i['impact'] == 'High':
                threat_intel_summary_data[2][13] = i['count']
            elif i['impact'] == 'Critical':
                threat_intel_summary_data[1][13] = i['count']
           
        total_closed_by_priority = 0
        total_pending_by_priority = 0
       
        priorities = ["P1", "P2", "P3", "P4"]
        count = 1
        for priority in priorities:
            closed_with_resolution = data_third["closed_with_resolution"][priority] + data_v1["closed_with_resolution"][priority]
            total_closed_by_priority += closed_with_resolution
           
            closed_without_acknowledgement = data_third["closed_without_acknowledgement"][priority] + data_v1["closed_without_acknowledgement"][priority]
            total_closed_by_priority += closed_without_acknowledgement
           
            pending_with_soc_team = data_third["pending_with_soc_team"][priority] + data_v1["pending_with_soc_team"][priority]
            total_pending_by_priority += pending_with_soc_team
           
            threat_intel_summary_data[count].extend([priority, closed_with_resolution + closed_without_acknowledgement, pending_with_soc_team, "", "", ""])
            db_data["Incident Summary by priority"][priority].update({"Closed Incidents": closed_with_resolution + closed_without_acknowledgement, "Pending from SOC Team": pending_with_soc_team})
            count += 1
        db_data["Incident Summary by status"] = {"Closed Incidents": total_closed_by_priority, "Pending Incidents": total_pending_by_priority}
                   
           
        total_closed_by_category = 0
        total_pending_by_category = 0
        i = 1
        for category in categories:
            closed_with_resolution = data_by_category["closed_with_resolution"][category]
            total_closed_by_category += closed_with_resolution
           
            closed_without_acknowledgement = data_by_category["closed_without_acknowledgement"][category]
            total_closed_by_category += closed_without_acknowledgement
           
            pending_with_soc_team = data_by_category["pending_with_soc_team"][category]
            total_pending_by_category += pending_with_soc_team
           
            if i <= len(priorities):
                threat_intel_summary_data[i].extend(["", category, closed_with_resolution + closed_without_acknowledgement, pending_with_soc_team, ""])
                i+= 1
            else:
                threat_intel_summary_data.append(["", "", "","", "","", "", "","", "", "","", "","", "", "","", "", "","", "","", "", "","", "", "","", "",  category, closed_with_resolution + closed_without_acknowledgement, pending_with_soc_team, ""])
            if category not in db_data["Incident Summary by category"]:
                db_data["Incident Summary by category"][category] = {}
            db_data["Incident Summary by category"][category].update({"Closed Incidents": closed_with_resolution + closed_without_acknowledgement, "Pending from SOC Team": pending_with_soc_team})
           
        visionOne_data["Threat Intel Summary"] = db_data
       
        threat_intel_summary_sheet.merge_cells('A1:C1')
        threat_intel_summary_sheet['A1'] = "Indicators of Compromise(IOC)"
       
        threat_intel_summary_sheet.merge_cells('G1:K1')
        threat_intel_summary_sheet['G1'] = "Matched IOCs Detailed Summary"
       
        threat_intel_summary_sheet.merge_cells('M1:N1')
        threat_intel_summary_sheet['M1'] = "Incident Summary by Severity"
       
        threat_intel_summary_sheet.merge_cells('AD1:AG1')
        threat_intel_summary_sheet['AD1'] = "Incidents_Summary_by_category"
       
        threat_intel_summary_sheet.merge_cells('W1:Z1')
        threat_intel_summary_sheet['W1'] = "Incidents_Summary_by_Priority"
 
        threat_intel_summary_sheet.merge_cells('R1:S1')
        threat_intel_summary_sheet['R1'] = "Incident Summary by Status"
       
        threat_intel_summary_sheet['E1'] = "R_S_N_IOC"
        threat_intel_summary_sheet['P1'] = "R_S_N_ISBS"
        threat_intel_summary_sheet['U1'] = "R_S_N_ISB_Status"
        threat_intel_summary_sheet['AB1'] = "R_S_N_ISBP"
        threat_intel_summary_sheet['AI1'] = "R_S_N_T10IS"
       
        for row in threat_intel_summary_data:
            threat_intel_summary_sheet.append(row)
       
        global total_closed_incidents
        total_closed_incidents = total_closed_by_priority
        threat_intel_summary_sheet["S3"] = total_closed_by_priority
        threat_intel_summary_sheet["S4"] = total_pending_by_priority
        threat_intel_summary_sheet['A7'] = "Email"
    except Exception as e:
        logging.error(f"Error Adding data to Overall Incidents Summary: {e}")
        return
                
###################################### Monthly and Weekly common Functions ####################################

def xl_adpotion_rate_a1():
    features = [
        ("eppAgentVulnerabilityProtection", "Vulnerability Protection"),
        ("eppAgentAntiMalwareScans", "Anti-Malware Scans"),
        ("eppAgentWebReputation", "Web Reputation"),
        ("eppAgentBehaviorMonitoring", "Behavior Monitoring"),
        ("eppAgentPredictiveMachineLearning", "Predictive Machine Learning"),
        ("eppAgentSmartFeedback", "Smart Feedback"),
        ("eppAgentFirewall", "Firewall"),
        ("eppAgentSuspiciousConnectionSettings", "Suspicious Connection Service")
    ]

    if is_monthly_report:
        features.extend([("eppAgentDeviceControl", "Device Control"), ("eppAgentApplicationControl", "Application Control")])
        
    data = []
    db_data = {"Key Feature Adoption Rate of A1": []}

    for feature, name in features:
        total = get_endpoint_details(f"{feature} eq enabled or {feature} eq disabled")
        count = get_endpoint_details(f"{feature} eq enabled")
        rate = round(count / total * 100, 1) if total > 0 else 0
        data.append([name, total, count, rate])
        db_data["Key Feature Adoption Rate of A1"].append({'Feature Name': name, 'Total': total, 'Count': count, 'Rate': rate})
        
    visionOne_data["Key Feature Adoption Rate of A1"] = db_data    
    try:
        key_feature_adoption_rate_of_a1_sheet = workbook.create_sheet(title="Key Feature Adoption Rate of A1")

        key_feature_adoption_rate_of_a1_sheet.merge_cells('A1:D1')
        key_feature_adoption_rate_of_a1_sheet['A1'] = "Key Feature Adoption Rate of A1"
        key_feature_adoption_rate_of_a1_sheet['F1'] = "Recommendation/Note/Summary"

        key_feature_adoption_rate_of_a1_sheet.append(["Feature's Names", "Total", "Count", "Rate"])
        
        for row in data:
            key_feature_adoption_rate_of_a1_sheet.append(row)
    except Exception as e:
        logging.error(f"Error Adding data to Key Feature Adoption Rate of A1: {e}")

def xl_adoption_rate_of_c1():
    c1_data = key_feature_adoption_rate_of_c1()
    try:
        key_feature_adoption_rate_of_c1_sheet = workbook.create_sheet(title="Key Feature Adoption Rate of C1")

        features = ["Intrusion Prevention System (IPS)", "Anti-malware", "Web Reputation", "Behavior Monitoring", "Predictive Machine Learning", "Smart Feedback", "Firewall", "Agent Self-Protection"]
        if is_monthly_report:
            features.extend(["File Integrity Monitoring (FIM)", "Log Inspection", "Device Control", "Application Control"])
            
        key_feature_adoption_rate_of_c1_data = [
            ["Feature's Names", "Total", "Count", "Rate"]
        ]

        db_data = {"Key Feature Adoption Rate of C1": []}
        for feature_name in features:
            total = c1_data.get(feature_name, {}).get("total", 'NA')
            count = c1_data.get(feature_name, {}).get("count", 'NA')
            if total != 'NA' and count != 'NA':
                if total == 0:
                    rate = 'NA'
                else:
                    rate = round(count / total*100)
            else:
                rate = 'NA'
            key_feature_adoption_rate_of_c1_data.append([feature_name, total, count, rate])
            db_data["Key Feature Adoption Rate of C1"].append({"Feature Name": feature_name, "Total": str(total), "Count": str(count), "Rate": str(rate)})
                    
        visionOne_data["Key Feature Adoption Rate of C1"] = db_data
        
        key_feature_adoption_rate_of_c1_sheet.merge_cells('A1:D1')
        key_feature_adoption_rate_of_c1_sheet['A1'] = "Key Feature Adoption Rate of C1WS"

        key_feature_adoption_rate_of_c1_sheet['F1'] = "Recommendation/Note/Summary"

        for row in key_feature_adoption_rate_of_c1_data:
            key_feature_adoption_rate_of_c1_sheet.append(row)
    except Exception as e:
        logging.error(f"Error Adding data to Key Feature Adoption Rate of C1: {e}")
        return
                                  
################################################ Main funtion #################################################
                                             
if __name__ == '__main__':
    logging.info("Script initiated")    
    read_config()

    parser = argparse.ArgumentParser(description="Processes tenant data with date and configuration validation")

    parser.add_argument('--date', type=str, required=True, help="Date in DD/MM/YYYY format for Weekly Reports to get the past one week Report from the given date or MM/YYYY format for Monthly Reports")
    parser.add_argument('--tenants', type=str, nargs='+', required=True, help="List of tenant names separated by space or 'all' to select all the tenants.")

    args = parser.parse_args()
    
    is_monthly_report, start_date, end_date, month_name, day = get_start_and_end_date(args.date)
   
    selected_tenant_list = get_tenant_name()
     
    ############### DB Connection #################
    
    connection_check, es = connect_elasticsearch()
    if connection_check:
        saved_reports = get_saved_reports_from_es()
    
    ############## Looping through required Tenant details #################
      
    for tenant_name in selected_tenant_list:
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)
        
        v1_base_url = config.get(tenant_name, 'v1_base_url')
        v1_token = config.get(tenant_name, 'v1_api_key')
        
        a1_base_url = config.get(tenant_name, 'a1_base_url')
        a1_application_id = config.get(tenant_name, 'a1_application_id')
        a1_api_key = config.get(tenant_name, 'a1_api_key')
        
        soar_base_url = config.get(tenant_name, 'soar_base_url')
        soar_api_key = config.get(tenant_name,'soar_api_key')
        
        cloud_app_sec_base_url = config.get(tenant_name, 'cloud_app_sec_base_url')
        cloud_app_sec_api_key = config.get(tenant_name, 'cloud_app_sec_api_key')
        
        c1_base_url = config.get(tenant_name, 'c1_base_url')
        c1_token = config.get(tenant_name, 'c1_api_key')
        
        headers = {
            'Authorization': 'Bearer ' + v1_token
        }
        
        isvalid_token()

        try:
            visionOne_data = {}
            wb_data = make_v1_api_call("/v3.0/workbench/alerts", {"top": 1000, "startDateTime": start_date, "endDateTime": end_date}, {})
            security_posture_data = get_security_posture()  

            ###################### Calling xl functions #################
            xl_adpotion_rate_a1()
            xl_adoption_rate_of_c1()
            if is_monthly_report:
                xl_top_risk_users()
                xl_top_risk_devices()
                xl_account_compromise_events()
                xl_critical_high_incidents_summary()
                xl_executive_summary()
                xl_overall_incidents_summary()
                xl_detailed_summary()
                xl_vulnerability_assessment_report()
                xl_system_configuration_report()
                xl_agent_version_summary()
                xl_detection_summary_from_a1() 
                xl_slo_summary()  
                xl_V1_workbench_incidents_summary()  
                xl_siem_incidents_summary()
                xl_pending_incidents_summary()
                xl_endpoint_feature_complicance()
                xl_email_summary()
                xl_risk_matrics()
                xl_top_vulnerability_detected()
                xl_product_assessment_report()
                get_component_versions()                
                file_name = f"{tenant_name}_{month_name}_Report"
                index = 'monthly_reports'  

            else:
                xl_weekly_threat_intel_summary()
                xl_weekly_executive_summary(start_date, end_date)
                xl_weekly_slo_summary()
                xl_weekly_pending_incidents_summary()
                xl_weekly_endpoint_inventory()
                file_name = f"{tenant_name}_{day}_{month_name}_Report"
                index = 'weekly_reports' 
            
            ####################### Saving Data ElasticSearch and Excel #################################
            if connection_check:
                is_file_present, file_id, document = get_id_of_es_file(file_name)
                print(file_name, 'out side')
                if is_file_present:
                    print('Overwriting already existing file')
                    response = es.index(index=index, id=file_id, document={'doc_name': file_name, 'document': visionOne_data})
                else:    
                    print('creating new file')
                    response = es.index(index=index, document={'doc_name': file_name, 'document': visionOne_data})

            workbook.save(f"{file_name}.xlsx")      
            
            # save_to_file("visionOne_data", visionOne_data)
        except Exception as e:
            logging.error(f"Unexpected error: {e}")
            
    logging.info("Script completed")  
    sys.exit(0)